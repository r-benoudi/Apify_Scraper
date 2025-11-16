from flask import Blueprint, render_template, jsonify, send_file
from app import db
# from app.models import MangaScraper
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime
import time

scraper_bp = Blueprint('scraper', __name__)

EXCEL_FILE = 'instance/manga_scraped.xlsx'

def fetch_data(url, headers):
    retries = 3
    for i in range(retries):
        try:
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                return response
            else:
                print(f"Erreur {response.status_code}, tentative {i+1}")
        except requests.exceptions.ConnectionError as e:
            print(f"Erreur de connexion, tentative {i+1}: {e}")
            time.sleep(3)  # Attendre 5 secondes avant de réessayer
    return None

def scrape_manga_page(page_num):
    """Scrape une page de manga"""
    try:
        url = f"https://lekmanga.net/manga/page/{page_num}"
        headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36 OPR/123.0.0.0"
        }
        
        # # Créer une session pour maintenir les cookies
        # session = requests.Session()
        
        # # Première requête pour obtenir les cookies
        # response = session.get(url, headers=headers, timeout=15, allow_redirects=True)
        # response.raise_for_status()
        response = fetch_data(url, headers=headers)
        
        soup = BeautifulSoup(response.content, 'html.parser')
        mangas = []
        
        # Trouver tous les items de manga
        manga_items = soup.find_all('div', class_='page-item-detail')

        print(f"Page {page_num}: {len(manga_items)} items trouvés")

        
        for item in manga_items:
            try:
                # Titre
                title_tag = item.find('h3', class_='h5')
                title = title_tag.find('a').text.strip() if title_tag and title_tag.find('a') else 'N/A'
                
                # Lien
                link = title_tag.find('a')['href'] if title_tag and title_tag.find('a') else 'N/A'
                
                # Rating (votes)
                rating_div = item.find('div', class_='post-total-rating')
                rating = rating_div.find('span', class_='score').text.strip() if rating_div else 'N/A'
                
                # Chapitres
                chapters = []
                chapter_items = item.find_all('div', class_='chapter-item')
                for ch in chapter_items[:2]:  # Seulement les 2 derniers
                    chapter_link = ch.find('span', class_='chapter')
                    chapter_date = ch.find('span', class_='post-on')
                    
                    if chapter_link and chapter_date:
                        chapters.append({
                            'title': chapter_link.find('a').text.strip() if chapter_link.find('a') else 'N/A',
                            'date': chapter_date.text.strip()
                        })
                
                mangas.append({
                    'title': title,
                    'link': link,
                    'rating': rating,
                    'chapter1_title': chapters[0]['title'] if len(chapters) > 0 else 'N/A',
                    'chapter1_date': chapters[0]['date'] if len(chapters) > 0 else 'N/A',
                    'chapter2_title': chapters[1]['title'] if len(chapters) > 1 else 'N/A',
                    'chapter2_date': chapters[1]['date'] if len(chapters) > 1 else 'N/A',
                    'page': page_num
                })
            except Exception as e:
                print(f"Erreur item: {e}")
                continue
        
        return mangas
    except requests.exceptions.HTTPError as e:
        print(f"Erreur HTTP page {page_num}: {e}")
        return []
    except Exception as e:
        print(f"Erreur page {page_num}: {e}")
        return []

def save_to_excel(all_mangas):
    """Sauvegarde les mangas dans un fichier Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mangas"
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['ID', 'Titre', 'Lien', 'Rating', 'Chapitre 1', 'Date Ch1', 'Chapitre 2', 'Date Ch2', 'Page', 'Date Scraping']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for idx, manga in enumerate(all_mangas, 1):
        ws.append([
            idx,
            manga['title'],
            manga['link'],
            manga['rating'],
            manga['chapter1_title'],
            manga['chapter1_date'],
            manga['chapter2_title'],
            manga['chapter2_date'],
            manga['page'],
            datetime.now().strftime('%Y-%m-%d %H:%M')
        ])
    
    # Ajuster largeur colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(EXCEL_FILE)

def load_from_excel():
    """Charge les mangas depuis le fichier Excel"""
    if not os.path.exists(EXCEL_FILE):
        return []
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    mangas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:  # Si titre existe
            mangas.append({
                'id': row[0],
                'title': row[1],
                'link': row[2],
                'rating': row[3],
                'chapter1_title': row[4],
                'chapter1_date': row[5],
                'chapter2_title': row[6],
                'chapter2_date': row[7],
                'page': row[8],
                'scraped_at': row[9]
            })
    
    return mangas

@scraper_bp.route('/scraper')
def scraper_page():
    """Page principale du scraper"""
    mangas = load_from_excel()
    return render_template('scraper.html', mangas=mangas, total=len(mangas))

@scraper_bp.route('/scraper/start', methods=['POST'])
def start_scraping():
    """Démarre le scraping"""
    try:
        all_mangas = []
        
        # IMPORTANT: Pour scraper TOUTES les 1892 pages, remplacer range(1, 6) par range(1, 1893)
        # Pour la démo, on fait seulement 10 pages pour tester
        for page in range(1, 3):  # Changez 3 en 1893 pour tout scraper
            mangas = scrape_manga_page(page)
            all_mangas.extend(mangas)
            time.sleep(3)  # Pause de 2 secondes entre chaque page pour respecter le serveur
            
            # Log progress
            print(f"Page {page} scrapée: {len(mangas)} mangas trouvés")
        
        if all_mangas:
            save_to_excel(all_mangas)
            return jsonify({
                'success': True,
                'message': f'{len(all_mangas)} mangas récupérés',
                'total': len(all_mangas)
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Aucun manga trouvé'
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erreur: {str(e)}'
        }), 500

@scraper_bp.route('/scraper/load')
def load_scraped():
    """Charge les mangas depuis Excel"""
    try:
        mangas = load_from_excel()
        return jsonify({
            'success': True,
            'mangas': mangas,
            'total': len(mangas)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erreur: {str(e)}'
        }), 500

@scraper_bp.route('/scraper/export/<format>')
def export_mangas(format):
    """Exporte les mangas"""
    items = load_from_excel()
    
    if format == 'csv':
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(['ID', 'Titre', 'Lien', 'Rating', 'Chapitre 1', 'Date Ch1', 'Chapitre 2', 'Date Ch2', 'Page'])
        for item in items:
            writer.writerow([
                item['id'], item['title'], item['link'], item['rating'],
                item['chapter1_title'], item['chapter1_date'],
                item['chapter2_title'], item['chapter2_date'], item['page']
            ])
        output = io.BytesIO()
        output.write(si.getvalue().encode('utf-8'))
        output.seek(0)
        return send_file(output, mimetype='text/csv', as_attachment=True, 
                        download_name=f'mangas_{datetime.now().strftime("%Y%m%d")}.csv')
    else:
        return jsonify(items)

@scraper_bp.route('/scraper/import-to-db', methods=['POST'])
def import_to_database():
    """Importe les mangas scrapés dans la base de données"""
    try:
        from app.models import Manga
        
        mangas = load_from_excel()
        imported = 0
        
        for manga_data in mangas:
            # Vérifier si existe déjà
            existing = Manga.query.filter_by(title=manga_data['title']).first()
            if not existing:
                manga = Manga(
                    title=manga_data['title'],
                    author=manga_data['author'],
                    status=manga_data['status']
                )
                db.session.add(manga)
                imported += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{imported} mangas importés dans la base de données'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Erreur: {str(e)}'
        }), 500
